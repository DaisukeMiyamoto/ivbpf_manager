#!/usr/bin/env python
# -*- coding: utf-8 -*-
from CosmoAPIClient import CosmoAPIClient
from xml.etree import ElementTree
import openpyxl as opx
import os


class XnpExcel:
    """
    Export IVBPF (CosmoDB) data to Excel file for excel2xoonips
    """

    def __init__(self):
        self.wb = opx.Workbook()
        self.ws = self.wb.active
        self.ws.title = 'CNS_Data'

        self.write_header()
        self.record_index = 2

    def write_header(self):
        header_data = ['ID',
                       '言語',
                       'タイトル',
                       'フリーキーワード',
                       '概要',
                       '日付',
                       'データタイプ',
                       '実験者',
                       'プレビューファイルパス',
                       'データファイルパス',
                       'ダウンロード制限',
                       'ダウンロード通知',
                       'README',
                       'Rights',
                       'インデックス']

        header_cns = ['タイトル',
                      '作成者',
                      'DOI',
                      '日付',
                      '言語',
                      'フリーキーワード',
                      '概要',
                      'プレビュー',
                      '画像',
                      'ファイル',
                      'Rights',
                      'URL',
                      'インデックス',
                      'ダウンロード制限',
                      'ダウンロード通知']

        header = header_cns

        for i, name in enumerate(header):
            self.ws.cell(column=i + 1, row=1, value=name)

    def add_record(self, record):
        item_list_data = ['doi',
                          'langs',
                          'title',
                          'keywords',
                          'description',
                          'date',
                          'data_type',
                          'experimenters',
                          'preview',
                          'data_files',
                          'download_limitation',
                          'download_notification',
                          'readme',
                          'rights',
                          'index']

        item_list_cns = ['title',
                         'creators',
                         'doi',
                         'date',
                         'langs',
                         'keywords',
                         'description',
                         'preview',
                         'images',
                         'data_files',
                         'rights',
                         'url',
                         'index',
                         'download_limitation',
                         'download_notification']

        item_list = item_list_cns

        record_fix = []
        for item in item_list:
            if item in record:
                record_fix.append(record[item])
            else:
                record_fix.append('')

        print(record_fix)

        for i, name in enumerate(record_fix):
            self.ws.cell(column=i + 1, row=self.record_index, value=name)

        self.record_index += 1

    def save(self, xls_filename):
        self.wb.save(filename=xls_filename)


if __name__ == '__main__':

    def detail2record(capi, data_id, settings, download_files=True):
        detailxml = capi.get_detail(data_id)

        # print detailxml
        detail_root = ElementTree.fromstring(detailxml.encode('utf-8'))

        # label
        title = detail_root[0].find('.//label').text
        if title is None or title == ' ---':
            title = settings['db_title'] + str(detail_root[0].attrib['data_id'])

        # description
        description = u''
        if detail_root[0].find('.//comment') is not None and detail_root[0].find('.//comment').text is not None:
            description = detail_root[0].find('.//comment').text.rstrip()

        # file/thumbnail
        thumbnail_file = u''
        image_list = []
        file_list = []
        for thumbnail in detail_root[0].find('.//thumbnails'):
            dirname = os.path.join('.', settings['local_file_dir'], str(detail_root[0].attrib['data_id']))
            subdirname = thumbnail.text.split('/')[-2]

            if not os.path.isdir(dirname):
                os.makedirs(dirname)

            if subdirname == settings['thumbnail_dir']:
                thumbnail_file = os.path.join(dirname, thumbnail.text.split('/')[-1])
                if download_files:
                    capi.get_file(thumbnail.text, thumbnail_file)

            else:
                if not os.path.isdir(os.path.join(dirname, subdirname)):
                    os.makedirs(os.path.join(dirname, subdirname))

                filename = os.path.join(dirname, subdirname, thumbnail.text.split('/')[-1])
                image_list.append(filename)
                if download_files:
                    capi.get_file(thumbnail.text, filename)

        image_list = list(set(image_list))
        images = u''
        for image in image_list:
            images += image + '\n'

        # files
        for item in detail_root[0].find('.//items').findall('.//item'):
            if item.attrib['type'] == 'file':
                subpath = ''
                if item.attrib['path'] != '':
                    subpath = '/' + item.attrib['path']

                file_list.append('file://'
                                 + settings['server_file_dir']
                                 + '/' + str(detail_root[0].attrib['data_id'])
                                 + '/data'
                                 + subpath
                                 + '/' + item.text)
                # print(file_list[-1])

        file_list = list(set(file_list))
        filepath = u''
        for path in file_list:
            filepath += path + '\n'

        # experimenters
        experimenters = detail_root[0].find('.//author').text
        if experimenters is None:
            experimenters = 'no data'

        # keywords
        keyword_list = []
        keyword_all = u''
        for metadata in detail_root[0].findall('.//component'):
            if metadata.text is not None and metadata.text != '':
                keyword_list.append(metadata.attrib['name'] + ': ' + metadata.text.replace('\n', ''))

        for keyword in detail_root[0].find('.//keywords'):
            keyword_list.append(keyword.text)

        keyword_list = list(set(keyword_list))
        for keyword in keyword_list:
            keyword_all += keyword + '\n'

        # indexes
        index_list = []
        indexes = u''
        index_list.append(settings['db_title'])
        for keyword in detail_root[0].find('.//keywords'):
            keyword_sep = keyword.text.split('/')
            for i in range(1, len(keyword_sep)):
                joined = ''.join(keyword_sep[0:i])
                index_list.append(settings['db_title'] + '/Keyword/' + joined)

            index_list.append(settings['db_title'] + '/Keyword/' + keyword.text)

        index_list = list(set(index_list))

        for index in index_list:
            indexes += '/Private/' + index + '\n' + '/Public/' + index + '\n'

        # url
        url_list = []
        urls = ''
        url_list.append(detail_root[0].attrib['url'])
        for link in detail_root[0].find('.//links').findall('.//link'):
            url_list.append(link.attrib['href'])
            if link.attrib['name'] is not None:
                pass

        for url in url_list:
            urls += url + '\n'

        print(urls)

        # summarize
        record = {'title': title,
                  'doi': settings['doi_prefix'] + str(detail_root[0].attrib['data_id']),
                  'keywords': keyword_all.rstrip('\n'),
                  'description': description,
                  'date': detail_root[0].find('.//date').text.replace('-', '/'),
                  'data_type': 'other',
                  'creators': experimenters,
                  'preview': thumbnail_file.replace('\\', '/').rstrip('\n'),
                  'images': images.rstrip('\n'),
                  'data_files': filepath.replace('\\', '/').rstrip('\n'),
                  'download_limitation': 'FALSE',
                  'download_notification': 'FALSE',
                  'readme': 'README',
                  'url': urls.rstrip('\n'),
                  'rights': 'CC-BY',
                  'index': (indexes + settings['additional_indexes']).rstrip('\n')
                  }

        print(record)

        return record


    import db_settings

    # choose settings
    settings = db_settings.settings_newdb1
    # settings = settings_newdb2

    capi = CosmoAPIClient(settings['url'], settings['db_name'], debug=False)
    list_xml = capi.get_list()

    root = ElementTree.fromstring(list_xml.encode('utf-8'))

    record_list = []
    thumbnail_list = []

    for child in root[1]:
        record = detail2record(capi, int(child.attrib['data_id']), settings, download_files=True)
        record_list.append(record)

    xnpxls = XnpExcel()

    for record in record_list:
        xnpxls.add_record(record)

    xnpxls.save(settings['output_filename'])
